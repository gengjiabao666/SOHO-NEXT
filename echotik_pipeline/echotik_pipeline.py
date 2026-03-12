#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
EchoTik 一键流水线（raw -> clean -> candidates）
用法（示例）：
  conda activate echotik_clean
  python /home/gjb/workspace/echotik_pipeline.py --captured 2026-03-03

默认：
- captured 默认今天（系统日期）
- geo 默认 US
- top 默认 200
- period 默认自动推断：
  - d: captured-1 天 的 MMDD（例如 0303 -> 0302）
  - w: captured 的上周周一-周日（MMDD-MMDD）
  - m: captured 的上个月 YYYYMM

你只要把 EchoTik 导出的 xlsx 放到 inbox/d|w|m 中即可。
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path
from datetime import date, datetime, timedelta
from typing import Optional, Dict, List

import pandas as pd
import openpyxl


# -------------------------
# 基础工具：日期/period 推断
# -------------------------
def mmdd(d: date) -> str:
    return d.strftime("%m%d")

def yyyymm(d: date) -> str:
    return d.strftime("%Y%m")

def yesterday(d: date) -> str:
    return mmdd(d - timedelta(days=1))

def last_week_range(d: date) -> str:
    """
    以 captured 日期为锚点，取“上周一~上周日”
    """
    # 本周一
    this_monday = d - timedelta(days=d.weekday())
    last_monday = this_monday - timedelta(days=7)
    last_sunday = this_monday - timedelta(days=1)
    return f"{mmdd(last_monday)}-{mmdd(last_sunday)}"

def last_month(d: date) -> str:
    """
    上个月 YYYYMM
    """
    first = d.replace(day=1)
    prev_last = first - timedelta(days=1)
    return yyyymm(prev_last)

def parse_date_yyyy_mm_dd(s: str) -> date:
    return datetime.strptime(s, "%Y-%m-%d").date()


# -------------------------
# 识别 ds：商品(p) / 新品(pn) / 小店(s)
# -------------------------
def detect_ds_by_header(xlsx_path: Path) -> Optional[str]:
    """
    通过表头和文件名识别：
    - 商品表（热销榜）：含 商品Id/商品ID/商品id/Product Id，文件名不含 "New"
    - 商品表（新品榜）：含 商品Id/商品ID/商品id/Product Id，文件名含 "New"
    - 小店表：含 店铺名称/Shop Name
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    is_product = False
    is_shop = False

    for r in range(1, 8):
        header = []
        for cell in list(ws[r])[:80]:
            if cell.value is None:
                continue
            header.append(str(cell.value).strip())
        text = " ".join(header)
        if ("商品Id" in text) or ("商品ID" in text) or ("商品id" in text) or ("Product Id" in text):
            is_product = True
            break
        if ("店铺名称" in text) or ("Shop Name" in text):
            is_shop = True
            break

    wb.close()

    if is_shop:
        return "s"
    if is_product:
        # 检查文件名是否含 "New" 或 "new"
        if "new" in xlsx_path.name.lower():
            return "pn"
        else:
            return "p"

    return None


# -------------------------
# 标准文件名生成
# et_{ds}_{win}_{period}_cap{cap}_n{top}_{geo}.xlsx
# -------------------------
def build_std_name(ds: str, win: str, period: str, cap_mmdd: str, top: str, geo: str, category: str = "") -> str:
    cat_suffix = f"_{category}" if category else ""
    return f"et_{ds}_{win}_{period}_cap{cap_mmdd}_n{top}_{geo}{cat_suffix}.xlsx"


# -------------------------
# 读取 xlsx -> DataFrame（表头行定位 + 行数上限）
# -------------------------
def find_header_row(ws, max_scan_rows: int = 12, max_scan_cols: int = 80) -> int:
    for r in range(1, max_scan_rows + 1):
        values = [ws.cell(r, c).value for c in range(1, max_scan_cols + 1)]
        text = " ".join([str(v).strip() for v in values if v is not None])
        if (("商品Id" in text) or ("商品ID" in text) or ("商品id" in text) or ("Product Id" in text)
                or ("店铺名称" in text) or ("Shop Name" in text)):
            return r
    raise RuntimeError("找不到表头行（未命中 商品Id/Product Id/店铺名称/Shop Name）")

def read_sheet_as_df(xlsx_path: Path, max_data_rows: int = 5000) -> pd.DataFrame:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    header_row = find_header_row(ws)

    headers = []
    for c in range(1, 80):
        v = ws.cell(header_row, c).value
        headers.append(str(v).strip() if v is not None else None)

    max_r = min(ws.max_row, header_row + max_data_rows)
    rows = []
    for r in range(header_row + 1, max_r + 1):
        row = []
        empty = True
        for c in range(1, 80):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip() != "":
                empty = False
            row.append(v)
        if empty:
            continue
        rows.append(row)

    wb.close()

    df = pd.DataFrame(rows, columns=headers)
    df = df.loc[:, [c for c in df.columns if c is not None]]
    return df


# -------------------------
# 小店：提取超链接 shop_url + echotik_shop_id
# -------------------------
SHOP_ID_RE = re.compile(r"/shops/(\d+)")

def extract_shop_url_and_id(xlsx_path: Path, df: pd.DataFrame) -> pd.DataFrame:
    """
    必须 read_only=False 才能取 hyperlink
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=False)
    ws = wb.active

    header_row = None
    col_shopname = None
    col_more = None

    for r in range(1, 13):
        for c in range(1, 81):
            v = ws.cell(r, c).value
            if v in ("店铺名称", "Shop Name"):
                header_row = r
                col_shopname = c
            elif v in ("查看更多", "View More"):
                header_row = header_row or r
                col_more = c
        if header_row and col_shopname:
            break

    if not header_row or not col_shopname:
        wb.close()
        df["echotik_shop_url"] = ""
        df["echotik_shop_id"] = ""
        return df

    urls: List[str] = []
    ids: List[str] = []

    start_r = header_row + 1
    for i in range(len(df)):
        r = start_r + i
        url = ""

        cell = ws.cell(r, col_shopname)
        if cell.hyperlink is not None and getattr(cell.hyperlink, "target", None):
            url = cell.hyperlink.target

        if (not url) and col_more:
            cell2 = ws.cell(r, col_more)
            if cell2.hyperlink is not None and getattr(cell2.hyperlink, "target", None):
                url = cell2.hyperlink.target

        urls.append(url or "")
        m = SHOP_ID_RE.search(url or "")
        ids.append(m.group(1) if m else "")

    wb.close()

    df["echotik_shop_url"] = urls
    df["echotik_shop_id"] = ids
    return df


# -------------------------
# clean v0（含 rank）
# -------------------------
GROUP_KEYS = ["dataset", "geo", "top_n", "window", "period"]

def add_ranks(df: pd.DataFrame) -> pd.DataFrame:
    """
    生成：
    - rank_gmv：按 gmv_usd 分组内排名（dense）
    - rank_units：按 units_sold 分组内排名（dense）
    - rank：默认 rank_gmv
    """
    df = df.copy()
    df["gmv_usd"] = pd.to_numeric(df.get("gmv_usd"), errors="coerce").fillna(0)
    df["units_sold"] = pd.to_numeric(df.get("units_sold"), errors="coerce").fillna(0)

    df["rank_gmv"] = (
        df.groupby(GROUP_KEYS)["gmv_usd"]
          .rank(method="dense", ascending=False)
          .astype("int64")
    )
    df["rank_units"] = (
        df.groupby(GROUP_KEYS)["units_sold"]
          .rank(method="dense", ascending=False)
          .astype("int64")
    )
    df["rank"] = df["rank_gmv"].astype("int64")
    return df


def normalize_products(df: pd.DataFrame, meta: dict) -> pd.DataFrame:
    rename_map = {}
    for c in df.columns:
        if c in ("商品Id", "商品ID", "商品id", "Product Id"):
            rename_map[c] = "echotik_product_id"
        elif c in ("商品名称", "Product Name"):
            rename_map[c] = "product_name"
        elif c in ("销量", "Sales"):
            rename_map[c] = "units_sold"
        elif c in ("销售额($)", "销售额", "GMV($)", "GMV"):
            rename_map[c] = "gmv_usd"
        elif c in ("带货分类", "类目", "分类", "Category"):
            rename_map[c] = "category"
        elif c in ("地区", "国家", "Region"):
            rename_map[c] = "region"

    df = df.rename(columns=rename_map)

    df["source"] = "echotik"
    # 根据 ds 区分数据集：p=热销榜，pn=新品榜
    if meta.get("ds") == "pn":
        df["dataset"] = "products_new"
    else:
        df["dataset"] = "products_hot"
    
    df["geo"] = meta["geo"]
    df["top_n"] = int(meta["top"])
    df["window"] = meta["win"]
    df["period"] = meta["period"]
    df["captured_mmdd"] = meta["cap"]

    out_cols = [
        "source","dataset","geo","top_n","window","period","captured_mmdd",
        "echotik_product_id","product_name","category","units_sold","gmv_usd"
    ]
    for c in out_cols:
        if c not in df.columns:
            df[c] = ""

    out = df[out_cols].copy()
    out["units_sold"] = pd.to_numeric(out["units_sold"], errors="coerce")
    out["gmv_usd"] = pd.to_numeric(out["gmv_usd"], errors="coerce")
    return out


def normalize_shops(df: pd.DataFrame, meta: dict, xlsx_path: Path) -> pd.DataFrame:
    df = extract_shop_url_and_id(xlsx_path, df)

    rename_map = {}
    for c in df.columns:
        if c in ("店铺名称", "Shop Name"):
            rename_map[c] = "shop_name"
        elif c in ("销量", "Sales"):
            rename_map[c] = "units_sold"
        elif c in ("销售额($)", "销售额", "GMV($)", "GMV"):
            rename_map[c] = "gmv_usd"
        elif c in ("带货分类", "类目", "分类", "Product Category"):
            rename_map[c] = "category"
        elif c in ("地区", "国家", "Region"):
            rename_map[c] = "region"

    df = df.rename(columns=rename_map)

    df["source"] = "echotik"
    df["dataset"] = "shops_crossborder"
    df["geo"] = meta["geo"]
    df["top_n"] = int(meta["top"])
    df["window"] = meta["win"]
    df["period"] = meta["period"]
    df["captured_mmdd"] = meta["cap"]

    out_cols = [
        "source","dataset","geo","top_n","window","period","captured_mmdd",
        "echotik_shop_id","shop_name","echotik_shop_url","category","units_sold","gmv_usd"
    ]
    for c in out_cols:
        if c not in df.columns:
            df[c] = ""

    out = df[out_cols].copy()
    out["units_sold"] = pd.to_numeric(out["units_sold"], errors="coerce")
    out["gmv_usd"] = pd.to_numeric(out["gmv_usd"], errors="coerce")
    return out


# -------------------------
# candidates v0（含 view）
# -------------------------
def top_n_by_metric(df: pd.DataFrame, metric: str, key_col: str, n: int = 50) -> pd.DataFrame:
    df = df.copy()
    df[metric] = pd.to_numeric(df.get(metric), errors="coerce").fillna(0)
    return df.sort_values([metric, key_col], ascending=[False, True]).head(n)

def ensure_cols(df: pd.DataFrame, cols: List[str]) -> pd.DataFrame:
    df = df.copy()
    for c in cols:
        if c not in df.columns:
            df[c] = ""
    return df[cols]


# -------------------------
# 主流程：inbox -> raw -> clean -> candidates
# -------------------------
def list_xlsx(folder: Path) -> List[Path]:
    if not folder.exists():
        return []
    return sorted(folder.rglob("*.xlsx"))


def main():
    ap = argparse.ArgumentParser(description="EchoTik pipeline: inbox -> raw -> clean -> candidates")
    ap.add_argument("--root", default="/mnt/g/SOHO_repo", help="SOHO_repo 根目录（WSL路径）")
    ap.add_argument("--captured", default="", help="采集日期 YYYY-MM-DD（默认今天）")
    ap.add_argument("--geo", default="US", help="地区（默认 US）")
    ap.add_argument("--top", default="200", help="Top N（默认 200）")
    ap.add_argument("--dry-run", action="store_true", help="只预演，不移动文件、不写输出")
    ap.add_argument("--daily", default="", help="日榜 period（MMDD），留空则自动 yesterday(captured)")
    ap.add_argument("--weekly", default="", help="周榜 period（MMDD-MMDD），留空则自动 last_week_range(captured)")
    ap.add_argument("--monthly", default="", help="月榜 period（YYYYMM），留空则自动 last_month(captured)")
    args = ap.parse_args()

    root = Path(args.root).expanduser().resolve()
    captured_date = parse_date_yyyy_mm_dd(args.captured) if args.captured else date.today()
    captured_str = captured_date.isoformat()
    cap_mmdd = mmdd(captured_date)

    # period 默认值
    daily_period = args.daily or yesterday(captured_date)
    weekly_period = args.weekly or last_week_range(captured_date)
    monthly_period = args.monthly or last_month(captured_date)
    period_map: Dict[str, str] = {"d": daily_period, "w": weekly_period, "m": monthly_period}

    # 路径
    base = root / "03_data_sources" / "echotik"
    inbox = base / "inbox"
    exports = base / "exports" / f"captured={captured_str}"
    raw_dir = exports / "raw"
    clean_dir = exports / "clean"
    cand_dir = exports / "candidates"

    # 创建目录
    for p in [raw_dir, clean_dir, cand_dir]:
        if args.dry_run:
            print(f"[DRY-RUN] mkdir -p {p}")
        else:
            p.mkdir(parents=True, exist_ok=True)

    print(f"[INFO] captured={captured_str} cap_mmdd={cap_mmdd} geo={args.geo} top={args.top}")
    print(f"[INFO] period: d={daily_period} w={weekly_period} m={monthly_period}")

    # 1) inbox -> raw（标准命名）
    moved_files: List[Path] = []
    for win in ["d", "w", "m"]:
        folder = inbox / win
        files = list_xlsx(folder)
        if not files:
            continue

        for f in files:
            ds = detect_ds_by_header(f)
            if ds is None:
                raise RuntimeError(f"无法识别 ds（商品/小店）：{f.name}，请确认表头含“商品id”或“店铺名称”")

            period = period_map[win]
            # 从路径提取品类（如 inbox/d/pet_supplies/xxx.xlsx -> pet_supplies）
            category = ""
            rel_parts = f.relative_to(folder).parts
            if len(rel_parts) > 1:
                category = rel_parts[0]  # 子目录名即品类
            new_name = build_std_name(ds, win, period, cap_mmdd, args.top, args.geo, category)
            dst = raw_dir / new_name

            if dst.exists():
                raise FileExistsError(f"目标文件已存在，避免覆盖：{dst}")

            if args.dry_run:
                print(f"[DRY-RUN] move {f} -> {dst}")
            else:
                f.rename(dst)
                moved_files.append(dst)
                print(f"[OK] {f.name} -> {dst.name}")

    # 2) raw -> clean（含 rank）
    raw_files = sorted(raw_dir.glob("et_*.xlsx"))
    if not raw_files:
        print("[WARN] raw 目录没有 et_*.xlsx（可能 inbox 为空，或 dry-run）")
        return

    products_frames = []
    shops_frames = []

    for f in raw_files:
        print(f"[INFO] clean processing {f.name}", flush=True)
        m = re.match(r"^et_(?P<ds>p|pn|s)_(?P<win>[dwm])_(?P<period>[^_]+)_cap(?P<cap>\d{4})_n(?P<top>\d+?)_(?P<geo>[A-Z]{2})(?:_(?P<category>[a-z_]+))?\.xlsx$", f.name)
        if not m:
            raise RuntimeError(f"raw 文件名不符合规范：{f.name}")
        meta = m.groupdict()

        df = read_sheet_as_df(f)
        meta2 = {"win": meta["win"], "period": meta["period"], "cap": meta["cap"], "top": meta["top"], "geo": meta["geo"], "ds": meta["ds"]}

        if meta["ds"] in ("p", "pn"):
            products_frames.append(normalize_products(df, meta2))
        else:
            shops_frames.append(normalize_shops(df, meta2, f))

    products = pd.concat(products_frames, ignore_index=True) if products_frames else pd.DataFrame()
    shops = pd.concat(shops_frames, ignore_index=True) if shops_frames else pd.DataFrame()

    if not products.empty:
        products = add_ranks(products)
    if not shops.empty:
        shops = add_ranks(shops)

    products_out = clean_dir / "products_clean_v0.csv"
    shops_out = clean_dir / "shops_clean_v0.csv"

    if args.dry_run:
        print(f"[DRY-RUN] write {products_out}")
        print(f"[DRY-RUN] write {shops_out}")
    else:
        # 固定列顺序
        p_cols = [
            "source","dataset","geo","top_n","window","period","captured_mmdd",
            "rank","rank_gmv","rank_units",
            "echotik_product_id","product_name","category","units_sold","gmv_usd"
        ]
        s_cols = [
            "source","dataset","geo","top_n","window","period","captured_mmdd",
            "rank","rank_gmv","rank_units",
            "echotik_shop_id","shop_name","echotik_shop_url","category","units_sold","gmv_usd"
        ]
        for c in p_cols:
            if c not in products.columns:
                products[c] = ""
        for c in s_cols:
            if c not in shops.columns:
                shops[c] = ""

        products[p_cols].to_csv(products_out, index=False, encoding="utf-8-sig")
        shops[s_cols].to_csv(shops_out, index=False, encoding="utf-8-sig")
        print(f"[OK] clean products -> {products_out} rows={len(products)}")
        print(f"[OK] clean shops -> {shops_out} rows={len(shops)}")

    # 3) clean -> candidates
    if args.dry_run:
        print("[DRY-RUN] candidates generation skipped")
        return

    products = pd.read_csv(products_out, dtype={"echotik_product_id": "string"})
    shops = pd.read_csv(shops_out, dtype={"echotik_shop_id": "string"})

    p_gmv = top_n_by_metric(products, "gmv_usd", "echotik_product_id", 50)
    p_units = top_n_by_metric(products, "units_sold", "echotik_product_id", 50)
    s_gmv = top_n_by_metric(shops, "gmv_usd", "echotik_shop_id", 50)
    s_units = top_n_by_metric(shops, "units_sold", "echotik_shop_id", 50)

    p_full_cols = [
        "source","dataset","geo","top_n","window","period","captured_mmdd",
        "rank","rank_gmv","rank_units",
        "echotik_product_id","product_name","category","units_sold","gmv_usd"
    ]
    s_full_cols = [
        "source","dataset","geo","top_n","window","period","captured_mmdd",
        "rank","rank_gmv","rank_units",
        "echotik_shop_id","shop_name","echotik_shop_url","category","units_sold","gmv_usd"
    ]

    (ensure_cols(p_gmv, p_full_cols)).to_csv(cand_dir / "candidate_products_v0_by_gmv_top50.csv", index=False, encoding="utf-8-sig")
    (ensure_cols(p_units, p_full_cols)).to_csv(cand_dir / "candidate_products_v0_by_units_top50.csv", index=False, encoding="utf-8-sig")
    (ensure_cols(s_gmv, s_full_cols)).to_csv(cand_dir / "candidate_shops_v0_by_gmv_top50.csv", index=False, encoding="utf-8-sig")
    (ensure_cols(s_units, s_full_cols)).to_csv(cand_dir / "candidate_shops_v0_by_units_top50.csv", index=False, encoding="utf-8-sig")

    # view 版（便于人工筛）
    p_view_cols = ["rank","window","period","product_name","category","units_sold","gmv_usd","echotik_product_id"]
    s_view_cols = ["rank","window","period","shop_name","category","units_sold","gmv_usd","echotik_shop_id","echotik_shop_url"]

    ensure_cols(p_gmv, p_view_cols).to_csv(cand_dir / "view_candidate_products_v0_by_gmv_top50.csv", index=False, encoding="utf-8-sig")
    ensure_cols(s_gmv, s_view_cols).to_csv(cand_dir / "view_candidate_shops_v0_by_gmv_top50.csv", index=False, encoding="utf-8-sig")

    print(f"[OK] candidates written to: {cand_dir}")

    # 4) manifest
    manifest = exports / "manifest.txt"
    if not manifest.exists():
        manifest.write_text(
            "\n".join([
                f"source: echotik",
                f"captured: {captured_str}",
                f"cap_mmdd: {cap_mmdd}",
                f"geo: {args.geo}",
                f"top_n: {args.top}",
                f"period_d: {daily_period}",
                f"period_w: {weekly_period}",
                f"period_m: {monthly_period}",
                f"notes: pipeline created raw/clean/candidates; shop_id extracted from hyperlink URL.",
            ]) + "\n",
            encoding="utf-8"
        )
        print(f"[OK] manifest -> {manifest}")
    else:
        print(f"[INFO] manifest exists, skip: {manifest}")


if __name__ == "__main__":
    main()